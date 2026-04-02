import customtkinter as ctk
import threading
import time
import random
import re
import os
import json
from datetime import datetime
from tkinter import filedialog, messagebox
import tkinter as tk
from tkinter import ttk

try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.chrome.service import Service
    from webdriver_manager.chrome import ChromeDriverManager
    SELENIUM_OK = True
except ImportError:
    SELENIUM_OK = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

# ─────────────────────────────────────────
#  Theme
# ─────────────────────────────────────────
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

BG_DARK   = "#0d1117"
BG_CARD   = "#161b22"
BG_INPUT  = "#21262d"
ACCENT    = "#58a6ff"
ACCENT2   = "#3fb950"
ACCENT3   = "#f78166"
TEXT_MAIN = "#e6edf3"
TEXT_DIM  = "#8b949e"
BORDER    = "#30363d"
ROW_ODD   = "#161b22"
ROW_EVEN  = "#1c2128"
HEADER_BG = "#1f6feb"


# ═══════════════════════════════════════════════════════
#  SCRAPER ENGINE  –  Phase 1: collect URLs, Phase 2: extract
# ═══════════════════════════════════════════════════════
class GoogleMapsScraper:
    def __init__(self, log_cb=None, progress_cb=None, result_cb=None):
        self.log      = log_cb      or print
        self.progress = progress_cb or (lambda pct, n: None)
        self.result   = result_cb   or (lambda d: None)   # called per record
        self.driver   = None
        self.stop_flag = False
        self.results  = []

    # ──────────────────────────────────────
    #  Driver
    # ──────────────────────────────────────
    def _init_driver(self, headless=True):
        opts = Options()
        if headless:
            opts.add_argument("--headless=new")
        opts.add_argument("--no-sandbox")
        opts.add_argument("--disable-dev-shm-usage")
        opts.add_argument("--disable-gpu")
        opts.add_argument("--window-size=1920,1080")
        opts.add_argument("--disable-blink-features=AutomationControlled")
        opts.add_experimental_option("excludeSwitches", ["enable-automation"])
        opts.add_experimental_option("useAutomationExtension", False)
        opts.add_argument(
            "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        )
        svc = Service(ChromeDriverManager().install())
        self.driver = webdriver.Chrome(service=svc, options=opts)
        self.driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
            "source": "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})"
        })

    def _wait_el(self, css, timeout=10):
        return WebDriverWait(self.driver, timeout).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, css))
        )

    def _els(self, css):
        try:
            return self.driver.find_elements(By.CSS_SELECTOR, css)
        except:
            return []

    def _el_text(self, css, default=""):
        els = self._els(css)
        return els[0].text.strip() if els else default

    def _attr(self, css, attr, default=""):
        els = self._els(css)
        return (els[0].get_attribute(attr) or "").strip() if els else default

    # ──────────────────────────────────────
    #  Phase 1 – Collect place URLs by scrolling
    # ──────────────────────────────────────
    def _collect_urls(self, keyword, max_results):
        self.log("🔍  Opening Google Maps search …")
        search_url = (
            "https://www.google.com/maps/search/"
            + keyword.replace(" ", "+")
        )
        self.driver.get(search_url)
        time.sleep(3)

        # consent pop-up
        for xp in [
            "//button[contains(.,'Accept all')]",
            "//button[contains(.,'Accept')]",
            "//button[contains(.,'Agree')]",
        ]:
            try:
                WebDriverWait(self.driver, 4).until(
                    EC.element_to_be_clickable((By.XPATH, xp))
                ).click()
                time.sleep(1)
                break
            except:
                pass

        # locate feed panel
        feed = None
        for css in [
            "div[role='feed']",
            "div.m6QErb[aria-label]",
            ".m6QErb.DxyBCb",
        ]:
            try:
                feed = WebDriverWait(self.driver, 8).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, css))
                )
                break
            except:
                pass

        if not feed:
            self.log("❌  Could not locate the results feed panel.")
            return []

        self.log("📜  Scrolling to collect place links …")
        urls      = []
        seen      = set()
        no_new    = 0
        max_no_new = 12   # stop if 12 consecutive scrolls bring nothing new

        while len(urls) < max_results and not self.stop_flag:
            # scrape all <a> tags that point to a /maps/place/
            anchors = self.driver.find_elements(
                By.CSS_SELECTOR,
                "a[href*='/maps/place/']"
            )
            added = 0
            for a in anchors:
                href = a.get_attribute("href") or ""
                # normalise – strip after the coordinate part
                m = re.match(r"(https://www\.google\.com/maps/place/[^?]+)", href)
                if m:
                    clean = m.group(1)
                else:
                    clean = href.split("?")[0]

                if clean and clean not in seen:
                    seen.add(clean)
                    urls.append(clean)
                    added += 1

            self.log(
                f"  ↕ Scrolled – {len(urls)} unique URLs collected …"
            )

            if added == 0:
                no_new += 1
            else:
                no_new = 0

            if no_new >= max_no_new:
                self.log("  ⚠ No new results after repeated scrolling – stopping scroll.")
                break

            # check for "end of results" sentinel
            end_divs = self.driver.find_elements(
                By.XPATH,
                "//*[contains(text(),'end of results') or "
                "contains(text(),\"You've reached\")]"
            )
            if end_divs:
                self.log("  ✅ Reached end of results.")
                break

            # scroll feed
            try:
                self.driver.execute_script(
                    "arguments[0].scrollTop += 2000", feed
                )
            except:
                self.driver.execute_script("window.scrollBy(0,1500)")
            time.sleep(random.uniform(1.2, 2.2))

        return urls[:max_results]

    # ──────────────────────────────────────
    #  Phase 2 – Visit each URL and extract
    # ──────────────────────────────────────
    def _extract_detail(self, url):
        data = {
            "Business Name" : "",
            "Category"      : "",
            "Rating"        : "",
            "Reviews"       : "",
            "Phone"         : "",
            "Email"         : "",
            "Website"       : "",
            "Address"       : "",
            "Hours"         : "",
            "Price Range"   : "",
            "Plus Code"     : "",
            "Contact Person": "",
            "Google Maps URL": url,
        }
        try:
            self.driver.get(url)
            WebDriverWait(self.driver, 15).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, "h1, h1.DUwDvf")
                )
            )
            time.sleep(random.uniform(1.5, 2.5))

            # ── Name
            for css in ["h1.DUwDvf", "h1[class*='fontHeadline']", "h1"]:
                t = self._el_text(css)
                if t:
                    data["Business Name"] = t
                    break

            # ── Category
            for css in [
                "button.DkEaL",
                "span.mgr77e",
                ".fontBodyMedium button[jsaction*='category']",
                "[jsaction*='pane.rating.category']",
            ]:
                t = self._el_text(css)
                if t:
                    data["Category"] = t
                    break

            # ── Rating
            for css in ["span.MW4etd", "div.F7nice span[aria-hidden='true']"]:
                t = self._el_text(css)
                if t:
                    data["Rating"] = t
                    break

            # ── Reviews
            for css in ["span.UY7F9", "div.F7nice span[aria-label]"]:
                els = self._els(css)
                if els:
                    raw = els[0].get_attribute("aria-label") or els[0].text
                    data["Reviews"] = re.sub(r"[^\d,]", "", raw)
                    break

            # ── Address  (button whose data-item-id contains "address")
            for css in [
                "button[data-item-id='address'] .Io6YTe",
                "[data-tooltip='Copy address'] .Io6YTe",
                "button[aria-label*='ddress'] .Io6YTe",
                ".rogA2c",                        # fallback
            ]:
                t = self._el_text(css)
                if t and len(t) > 5:
                    data["Address"] = t
                    break

            # ── Phone
            for css in [
                "button[data-item-id*='phone'] .Io6YTe",
                "[data-tooltip*='phone'] .Io6YTe",
                "button[aria-label*='hone'] .Io6YTe",
            ]:
                t = self._el_text(css)
                if t:
                    data["Phone"] = t
                    break
            if not data["Phone"]:
                # last resort: regex in page source
                src = self.driver.page_source
                m = re.search(r'(\+?[\d\s\-\(\)]{7,20})', src)
                if m:
                    data["Phone"] = m.group(1).strip()

            # ── Website
            for css in [
                "a[data-item-id='authority']",
                "a[aria-label*='ebsite']",
                "a[href*='http'][data-tooltip*='ebsite']",
            ]:
                els = self._els(css)
                if els:
                    href = els[0].get_attribute("href") or els[0].text
                    if href and "google.com" not in href:
                        data["Website"] = href
                        break

            # ── Hours
            for css in [
                "button[data-item-id*='oh'] .Io6YTe",
                ".t39EBf .ZDu9vd",
                "[aria-label*='hour'] .Io6YTe",
            ]:
                t = self._el_text(css)
                if t:
                    data["Hours"] = t
                    break

            # ── Price Range
            for css in ["span[aria-label*='rice']", "span.mgr77e"]:
                els = self._els(css)
                for el in els:
                    t = el.text.strip()
                    if t.startswith("$") or t.startswith("€") or t.startswith("£"):
                        data["Price Range"] = t
                        break

            # ── Plus Code
            t = self._el_text("button[data-item-id='oloc'] .Io6YTe")
            if t:
                data["Plus Code"] = t

            # ── Try email from website
            if data["Website"] and not data["Email"]:
                data["Email"] = self._fetch_email(data["Website"], url)

        except Exception as exc:
            self.log(f"    ⚠ extract error: {exc}")

        return data

    def _fetch_email(self, site_url, return_url):
        """Visit website, grab first e-mail found, come back."""
        try:
            if not site_url.startswith("http"):
                site_url = "https://" + site_url
            self.driver.get(site_url)
            time.sleep(random.uniform(1.5, 2.5))
            src = self.driver.page_source
            m = re.search(
                r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}', src
            )
            email = m.group(0) if m else ""

            # also check contact page
            if not email:
                for suffix in ["/contact", "/contact-us", "/about"]:
                    try:
                        self.driver.get(site_url.rstrip("/") + suffix)
                        time.sleep(1.5)
                        m2 = re.search(
                            r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}',
                            self.driver.page_source,
                        )
                        if m2:
                            email = m2.group(0)
                            break
                    except:
                        pass

            # restore map page
            self.driver.get(return_url)
            time.sleep(1)
            return email
        except:
            try:
                self.driver.get(return_url)
                time.sleep(1)
            except:
                pass
            return ""

    # ──────────────────────────────────────
    #  Main entry
    # ──────────────────────────────────────
    def scrape(self, keyword, max_results=100, headless=True):
        self.stop_flag = False
        self.results   = []

        self.log("🚀  Starting browser …")
        self._init_driver(headless)

        try:
            # ── Phase 1
            urls = self._collect_urls(keyword, max_results)
            if not urls:
                self.log("❌  No URLs found. Try a different keyword.")
                return []

            self.log(
                f"\n✅  Collected {len(urls)} place URLs. "
                "Starting detail extraction …\n"
                + "─" * 60
            )

            # ── Phase 2
            for idx, url in enumerate(urls, 1):
                if self.stop_flag:
                    break
                self.log(f"  [{idx}/{len(urls)}]  Visiting …")
                data = self._extract_detail(url)
                self.results.append(data)
                pct = int(idx / len(urls) * 100)
                self.progress(pct, idx)
                self.result(data)          # push to UI immediately
                self.log(
                    f"    ✅  {data['Business Name'] or '(no name)'}  |  "
                    f"📞 {data['Phone'] or '–'}  |  "
                    f"✉ {data['Email'] or '–'}  |  "
                    f"🌐 {data['Website'] or '–'}"
                )
                time.sleep(random.uniform(0.8, 1.8))

            self.log(
                f"\n🎉  Extraction complete – "
                f"{len(self.results)} records."
            )
            self.progress(100, len(self.results))
            return self.results

        finally:
            try:
                self.driver.quit()
            except:
                pass

    def stop(self):
        self.stop_flag = True
        try:
            self.driver.quit()
        except:
            pass


# ═══════════════════════════════════════════════════════
#  EXCEL EXPORTER
# ═══════════════════════════════════════════════════════
class ExcelExporter:
    COLS = [
        ("Business Name",  30),
        ("Category",       20),
        ("Rating",          8),
        ("Reviews",        10),
        ("Phone",          18),
        ("Email",          32),
        ("Website",        35),
        ("Address",        40),
        ("Hours",          22),
        ("Price Range",    12),
        ("Plus Code",      16),
        ("Contact Person", 20),
        ("Google Maps URL",50),
    ]

    def export(self, data, filepath, keyword=""):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Results"

        ncols = len(self.COLS)
        last_col = get_column_letter(ncols)

        # ── Title
        ws.merge_cells(f"A1:{last_col}1")
        tc = ws["A1"]
        tc.value = (
            f"Google Maps Extractor  |  Keyword: {keyword}  |  "
            f"Date: {datetime.now():%Y-%m-%d %H:%M}"
        )
        tc.font      = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
        tc.fill      = PatternFill(fill_type="solid", fgColor="1F6FEB")
        tc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 32

        # ── Headers
        thin   = Side(style="thin", color="30363D")
        bdr    = Border(left=thin, right=thin, top=thin, bottom=thin)
        h_fill = PatternFill(fill_type="solid", fgColor="0D1117")
        h_font = Font(name="Calibri", bold=True, size=11, color="58A6FF")
        h_aln  = Alignment(horizontal="center", vertical="center")

        for ci, (name, width) in enumerate(self.COLS, 1):
            c = ws.cell(row=2, column=ci, value=name)
            c.font = h_font; c.fill = h_fill
            c.alignment = h_aln; c.border = bdr
            ws.column_dimensions[get_column_letter(ci)].width = width
        ws.row_dimensions[2].height = 26

        # ── Data
        f_odd  = PatternFill(fill_type="solid", fgColor="161B22")
        f_even = PatternFill(fill_type="solid", fgColor="1C2128")
        d_font = Font(name="Calibri", size=10, color="E6EDF3")
        d_aln  = Alignment(vertical="center", wrap_text=False)

        for ri, rec in enumerate(data, 3):
            fill = f_odd if ri % 2 else f_even
            for ci, (name, _) in enumerate(self.COLS, 1):
                c = ws.cell(row=ri, column=ci, value=rec.get(name, ""))
                c.font = d_font; c.fill = fill
                c.alignment = d_aln; c.border = bdr
            ws.row_dimensions[ri].height = 18

        ws.freeze_panes = "A3"
        ws.auto_filter.ref = f"A2:{last_col}2"

        # ── Summary sheet
        ws2 = wb.create_sheet("Summary")
        rows = [
            ("Keyword",        keyword),
            ("Total Records",  len(data)),
            ("With Phone",     sum(1 for r in data if r.get("Phone"))),
            ("With Email",     sum(1 for r in data if r.get("Email"))),
            ("With Website",   sum(1 for r in data if r.get("Website"))),
            ("Exported At",    datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ]
        ws2["A1"].value = "Extraction Summary"
        ws2["A1"].font  = Font(bold=True, size=14, color="58A6FF")
        for i, (k, v) in enumerate(rows, 3):
            ws2[f"A{i}"] = k
            ws2[f"B{i}"] = v
            ws2[f"A{i}"].font = Font(bold=True, color="8B949E", size=11)
            ws2[f"B{i}"].font = Font(color="E6EDF3", size=11)
        ws2.column_dimensions["A"].width = 20
        ws2.column_dimensions["B"].width = 35

        wb.save(filepath)


# ═══════════════════════════════════════════════════════
#  MAIN UI
# ═══════════════════════════════════════════════════════
class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Google Maps Business Extractor")
        self.geometry("1340x860")
        self.minsize(1100, 720)
        self.configure(fg_color=BG_DARK)

        self.scraper = None
        self.results = []
        self._sort_reverse = {}

        self._build_ui()

    # ──────────────────────────────────────
    #  Top bar
    # ──────────────────────────────────────
    def _build_ui(self):
        self._topbar()
        body = ctk.CTkFrame(self, fg_color="transparent")
        body.pack(fill="both", expand=True, padx=14, pady=(0, 12))
        body.grid_columnconfigure(0, weight=0, minsize=310)
        body.grid_columnconfigure(1, weight=1)
        body.grid_rowconfigure(0, weight=1)
        self._left(body)
        self._right(body)

    def _topbar(self):
        bar = ctk.CTkFrame(self, fg_color=BG_CARD, corner_radius=0, height=56)
        bar.pack(fill="x")
        bar.pack_propagate(False)
        ctk.CTkLabel(
            bar,
            text="  🗺  Google Maps Business Extractor",
            font=ctk.CTkFont("Segoe UI", 20, "bold"),
            text_color=ACCENT,
        ).pack(side="left", padx=20)
        self.status_lbl = ctk.CTkLabel(
            bar, text="● Ready",
            font=ctk.CTkFont(size=12), text_color=ACCENT2,
            fg_color=BG_INPUT, corner_radius=10, padx=12, pady=3,
        )
        self.status_lbl.pack(side="right", padx=18)

    # ──────────────────────────────────────
    #  Left panel
    # ──────────────────────────────────────
    def _left(self, parent):
        f = ctk.CTkFrame(
            parent, fg_color=BG_CARD,
            corner_radius=12, border_width=1, border_color=BORDER
        )
        f.grid(row=0, column=0, sticky="nsew", padx=(0, 8), pady=8)

        self._sec(f, "🔍  Search Settings")

        self._lbl(f, "Search Keyword")
        self.kw_entry = ctk.CTkEntry(
            f, placeholder_text="e.g.  plumbers in Chicago",
            font=ctk.CTkFont(size=13), fg_color=BG_INPUT,
            border_color=BORDER, text_color=TEXT_MAIN,
            height=40, corner_radius=8,
        )
        self.kw_entry.pack(fill="x", padx=14, pady=(2, 10))
        self.kw_entry.bind("<Return>", lambda e: self._start())

        # max results entry (typed value, not slider)
        self._lbl(f, "Max Results  (e.g. 50, 200, 500, 1000)")
        self.max_entry = ctk.CTkEntry(
            f, placeholder_text="100",
            font=ctk.CTkFont(size=13), fg_color=BG_INPUT,
            border_color=BORDER, text_color=TEXT_MAIN,
            height=36, corner_radius=8, width=120,
        )
        self.max_entry.insert(0, "100")
        self.max_entry.pack(anchor="w", padx=14, pady=(2, 8))

        # headless
        self.headless_var = ctk.BooleanVar(value=True)
        ctk.CTkCheckBox(
            f, text="Run browser in background (headless)",
            variable=self.headless_var,
            font=ctk.CTkFont(size=11), text_color=TEXT_DIM,
            fg_color=ACCENT, checkmark_color="white",
        ).pack(anchor="w", padx=14, pady=(0, 6))

        # fetch email
        self.email_var = ctk.BooleanVar(value=True)
        ctk.CTkCheckBox(
            f, text="Try to fetch Email from website",
            variable=self.email_var,
            font=ctk.CTkFont(size=11), text_color=TEXT_DIM,
            fg_color=ACCENT, checkmark_color="white",
        ).pack(anchor="w", padx=14, pady=(0, 6))

        self._div(f)

        # ── Progress info
        self._sec(f, "📊  Live Progress")
        self.prog_bar = ctk.CTkProgressBar(
            f, height=14, corner_radius=7,
            fg_color=BG_INPUT, progress_color=ACCENT2,
        )
        self.prog_bar.set(0)
        self.prog_bar.pack(fill="x", padx=14, pady=(4, 2))
        self.prog_lbl = ctk.CTkLabel(
            f, text="0 / 0",
            font=ctk.CTkFont(size=11), text_color=TEXT_DIM,
        )
        self.prog_lbl.pack(pady=(0, 6))

        self.phase_lbl = ctk.CTkLabel(
            f, text="Phase: –",
            font=ctk.CTkFont(size=11, weight="bold"),
            text_color=ACCENT,
        )
        self.phase_lbl.pack(pady=(0, 8))

        self._div(f)

        # ── Stats
        self._sec(f, "📈  Stats")
        stats_f = ctk.CTkFrame(f, fg_color=BG_INPUT, corner_radius=8)
        stats_f.pack(fill="x", padx=14, pady=(0, 8))
        self.s_total  = self._chip(stats_f, "Total",    "0", ACCENT)
        self.s_phone  = self._chip(stats_f, "Phone",    "0", ACCENT2)
        self.s_email  = self._chip(stats_f, "Email",    "0", "#f0883e")
        self.s_web    = self._chip(stats_f, "Website",  "0", "#a371f7")

        self._div(f)

        # ── Buttons
        self.btn_start = ctk.CTkButton(
            f, text="▶  Start Extraction",
            font=ctk.CTkFont(size=14, weight="bold"),
            fg_color=ACCENT, hover_color="#1f6feb",
            height=44, corner_radius=10, command=self._start,
        )
        self.btn_start.pack(fill="x", padx=14, pady=(6, 4))

        self.btn_stop = ctk.CTkButton(
            f, text="⛔  Stop",
            font=ctk.CTkFont(size=13),
            fg_color=ACCENT3, hover_color="#b22222",
            height=38, corner_radius=10,
            state="disabled", command=self._stop,
        )
        self.btn_stop.pack(fill="x", padx=14, pady=4)

        self.btn_export = ctk.CTkButton(
            f, text="📊  Export to Excel",
            font=ctk.CTkFont(size=13),
            fg_color=ACCENT2, hover_color="#2ea043",
            height=38, corner_radius=10,
            state="disabled", command=self._export,
        )
        self.btn_export.pack(fill="x", padx=14, pady=4)

        ctk.CTkButton(
            f, text="🗑  Clear",
            font=ctk.CTkFont(size=12),
            fg_color=BG_INPUT, hover_color=BORDER,
            text_color=TEXT_DIM, height=32, corner_radius=10,
            command=self._clear,
        ).pack(fill="x", padx=14, pady=(4, 14))

    # ──────────────────────────────────────
    #  Right panel
    # ──────────────────────────────────────
    def _right(self, parent):
        f = ctk.CTkFrame(
            parent, fg_color=BG_CARD,
            corner_radius=12, border_width=1, border_color=BORDER
        )
        f.grid(row=0, column=1, sticky="nsew", pady=8)
        f.grid_rowconfigure(1, weight=1)
        f.grid_columnconfigure(0, weight=1)

        # Tab bar
        tb = ctk.CTkFrame(f, fg_color="transparent", height=44)
        tb.grid(row=0, column=0, sticky="ew", padx=14, pady=(12, 0))
        for text, key in [("📋  Results Table", "table"), ("📝  Log", "log")]:
            btn = ctk.CTkButton(
                tb, text=text,
                font=ctk.CTkFont(size=12, weight="bold"),
                fg_color=ACCENT if key == "table" else BG_INPUT,
                hover_color="#1f6feb",
                text_color=TEXT_MAIN if key == "table" else TEXT_DIM,
                width=160, height=32, corner_radius=8,
                command=lambda k=key: self._tab(k),
            )
            btn.pack(side="left", padx=(0, 6))
            setattr(self, f"tbtn_{key}", btn)

        # content
        cf = ctk.CTkFrame(f, fg_color="transparent")
        cf.grid(row=1, column=0, sticky="nsew", padx=14, pady=(8, 14))
        cf.grid_rowconfigure(0, weight=1)
        cf.grid_columnconfigure(0, weight=1)

        self._build_table(cf)
        self._build_log(cf)
        self._tab("table")

    def _build_table(self, parent):
        self.tbl_frame = ctk.CTkFrame(parent, fg_color="transparent")
        self.tbl_frame.grid(row=0, column=0, sticky="nsew")
        self.tbl_frame.grid_rowconfigure(0, weight=1)
        self.tbl_frame.grid_columnconfigure(0, weight=1)

        cols = (
            "#", "Business Name", "Category", "Rating",
            "Reviews", "Phone", "Email", "Website", "Address"
        )
        widths = [38, 200, 130, 58, 70, 140, 200, 200, 230]

        style = ttk.Style()
        style.theme_use("clam")
        style.configure(
            "G.Treeview",
            background=ROW_ODD, foreground=TEXT_MAIN,
            fieldbackground=ROW_ODD, rowheight=26,
            font=("Segoe UI", 10), borderwidth=0,
        )
        style.configure(
            "G.Treeview.Heading",
            background=HEADER_BG, foreground="white",
            font=("Segoe UI", 10, "bold"), borderwidth=0, relief="flat",
        )
        style.map("G.Treeview",
                  background=[("selected", "#1f6feb")],
                  foreground=[("selected", "white")])
        style.map("G.Treeview.Heading",
                  background=[("active", "#388bfd")])

        self.tree = ttk.Treeview(
            self.tbl_frame, columns=cols,
            show="headings", style="G.Treeview", selectmode="extended",
        )
        for col, w in zip(cols, widths):
            self.tree.heading(
                col, text=col,
                command=lambda c=col: self._sort(c)
            )
            self.tree.column(col, width=w, minwidth=30, stretch=False)

        self.tree.tag_configure("odd",  background=ROW_ODD)
        self.tree.tag_configure("even", background=ROW_EVEN)

        vsb = ttk.Scrollbar(
            self.tbl_frame, orient="vertical", command=self.tree.yview
        )
        hsb = ttk.Scrollbar(
            self.tbl_frame, orient="horizontal", command=self.tree.xview
        )
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        # context menu
        menu = tk.Menu(
            self, tearoff=0, bg=BG_CARD, fg=TEXT_MAIN,
            activebackground=ACCENT, activeforeground="white",
            font=("Segoe UI", 10),
        )
        menu.add_command(label="📋 Copy Row",       command=self._copy_row)
        menu.add_command(label="🗑 Delete Row",      command=self._del_row)
        menu.add_separator()
        menu.add_command(label="📊 Export Selected", command=self._export_sel)
        self.tree.bind("<Button-3>", lambda e: menu.tk_popup(e.x_root, e.y_root))

    def _build_log(self, parent):
        self.log_frame = ctk.CTkFrame(parent, fg_color="transparent")
        self.log_frame.grid(row=0, column=0, sticky="nsew")
        self.log_box = ctk.CTkTextbox(
            self.log_frame,
            font=ctk.CTkFont("Consolas", 11),
            fg_color=BG_INPUT, text_color="#c9d1d9",
            corner_radius=8, wrap="word",
        )
        self.log_box.pack(fill="both", expand=True)
        self.log_box.configure(state="disabled")

    # ──────────────────────────────────────
    #  Tab switch
    # ──────────────────────────────────────
    def _tab(self, key):
        self.tbl_frame.tkraise() if key == "table" else self.log_frame.tkraise()
        for k in ("table", "log"):
            b = getattr(self, f"tbtn_{k}")
            active = k == key
            b.configure(
                fg_color=ACCENT    if active else BG_INPUT,
                text_color=TEXT_MAIN if active else TEXT_DIM,
            )

    # ──────────────────────────────────────
    #  Helpers
    # ──────────────────────────────────────
    def _sec(self, p, t):
        ctk.CTkLabel(
            p, text=t,
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color=ACCENT, anchor="w",
        ).pack(fill="x", padx=14, pady=(12, 2))

    def _lbl(self, p, t):
        ctk.CTkLabel(
            p, text=t,
            font=ctk.CTkFont(size=10),
            text_color=TEXT_DIM, anchor="w",
        ).pack(fill="x", padx=14, pady=(4, 0))

    def _div(self, p):
        ctk.CTkFrame(p, height=1, fg_color=BORDER).pack(
            fill="x", padx=10, pady=6
        )

    def _chip(self, parent, label, val, color):
        f = ctk.CTkFrame(parent, fg_color="transparent")
        f.pack(side="left", padx=10, pady=6)
        v = ctk.CTkLabel(
            f, text=val,
            font=ctk.CTkFont(size=17, weight="bold"),
            text_color=color,
        )
        v.pack()
        ctk.CTkLabel(
            f, text=label,
            font=ctk.CTkFont(size=9),
            text_color=TEXT_DIM,
        ).pack()
        return v

    # ──────────────────────────────────────
    #  Logging
    # ──────────────────────────────────────
    def _log(self, msg):
        def _do():
            self.log_box.configure(state="normal")
            ts = datetime.now().strftime("%H:%M:%S")
            self.log_box.insert("end", f"[{ts}]  {msg}\n")
            self.log_box.see("end")
            self.log_box.configure(state="disabled")
        self.after(0, _do)

    def _log_clear(self):
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.configure(state="disabled")

    def _set_status(self, txt, color=ACCENT2):
        self.status_lbl.configure(text=f"● {txt}", text_color=color)

    def _set_phase(self, txt):
        self.after(0, lambda: self.phase_lbl.configure(text=f"Phase: {txt}"))

    # ──────────────────────────────────────
    #  Progress
    # ──────────────────────────────────────
    def _on_progress(self, pct, count):
        def _do():
            self.prog_bar.set(pct / 100)
            max_r = self._max_r
            self.prog_lbl.configure(text=f"{count} / {max_r}")
        self.after(0, _do)

    # ──────────────────────────────────────
    #  Per-record callback (live row insert)
    # ──────────────────────────────────────
    def _on_result(self, data):
        def _do():
            self.results.append(data)
            n = len(self.results)
            tag = "odd" if n % 2 else "even"
            self.tree.insert(
                "", "end", iid=str(n), tags=(tag,),
                values=(
                    n,
                    data.get("Business Name", ""),
                    data.get("Category", ""),
                    data.get("Rating", ""),
                    data.get("Reviews", ""),
                    data.get("Phone", ""),
                    data.get("Email", ""),
                    data.get("Website", ""),
                    data.get("Address", ""),
                ),
            )
            self.tree.see(str(n))
            # stats
            self.s_total.configure(text=str(n))
            self.s_phone.configure(
                text=str(sum(1 for r in self.results if r.get("Phone")))
            )
            self.s_email.configure(
                text=str(sum(1 for r in self.results if r.get("Email")))
            )
            self.s_web.configure(
                text=str(sum(1 for r in self.results if r.get("Website")))
            )
        self.after(0, _do)

    # ──────────────────────────────────────
    #  Start / Stop
    # ──────────────────────────────────────
    def _start(self):
        if not SELENIUM_OK:
            messagebox.showerror(
                "Missing",
                "pip install selenium webdriver-manager"
            )
            return
        kw = self.kw_entry.get().strip()
        if not kw:
            messagebox.showwarning("No keyword", "Enter a search keyword.")
            return
        try:
            max_r = int(self.max_entry.get().strip() or "100")
            if max_r < 1:
                raise ValueError
        except:
            messagebox.showwarning("Invalid", "Enter a valid number for Max Results.")
            return

        self._max_r  = max_r
        self.results = []
        self.tree.delete(*self.tree.get_children())
        self._log_clear()

        self.btn_start.configure(state="disabled")
        self.btn_stop.configure(state="normal")
        self.btn_export.configure(state="disabled")
        self.prog_bar.set(0)
        self.prog_lbl.configure(text=f"0 / {max_r}")
        self._set_status("⏳ Running …", "#f0883e")
        self._set_phase("Collecting URLs")
        self._log(f"Keyword  : {kw}")
        self._log(f"Max      : {max_r}")
        self._log("─" * 60)

        headless   = self.headless_var.get()
        fetch_email = self.email_var.get()

        self.scraper = GoogleMapsScraper(
            log_cb      = self._log,
            progress_cb = self._on_progress,
            result_cb   = self._on_result,
        )
        # patch: disable email fetch if checkbox off
        if not fetch_email:
            self.scraper._fetch_email = lambda *a, **kw: ""

        def _run():
            try:
                self.after(0, lambda: self._set_phase("Collecting URLs …"))
                res = self.scraper.scrape(kw, max_r, headless)
                self.after(0, lambda: self._done(res))
            except Exception as e:
                self.after(0, lambda: self._err(str(e)))

        threading.Thread(target=_run, daemon=True).start()

    def _stop(self):
        if self.scraper:
            self.scraper.stop()
        self._set_status("⛔ Stopped", ACCENT3)
        self.btn_start.configure(state="normal")
        self.btn_stop.configure(state="disabled")
        if self.results:
            self.btn_export.configure(state="normal")
        self._log("⛔ Stopped by user.")

    def _done(self, res):
        self._set_status(f"✅ Done – {len(res)} records", ACCENT2)
        self._set_phase("Complete ✅")
        self.btn_start.configure(state="normal")
        self.btn_stop.configure(state="disabled")
        if res:
            self.btn_export.configure(state="normal")
        self._tab("table")

    def _err(self, msg):
        self._log(f"❌ {msg}")
        self._set_status("❌ Error", ACCENT3)
        self.btn_start.configure(state="normal")
        self.btn_stop.configure(state="disabled")

    # ──────────────────────────────────────
    #  Export
    # ──────────────────────────────────────
    def _export(self):
        if not self.results:
            messagebox.showwarning("Empty", "No data to export.")
            return
        self._do_export(self.results)

    def _export_sel(self):
        sel = self.tree.selection()
        if not sel:
            return
        data = [self.results[int(i) - 1]
                for i in sel
                if 0 <= int(i) - 1 < len(self.results)]
        self._do_export(data)

    def _do_export(self, data):
        if not OPENPYXL_OK:
            messagebox.showerror("Missing", "pip install openpyxl")
            return
        kw  = self.kw_entry.get().strip()
        ts  = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe = re.sub(r'[\\/:*?"<>|]', "_", kw)[:40]
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"GoogleMaps_{safe}_{ts}.xlsx",
            title="Save Excel",
        )
        if not path:
            return
        try:
            ExcelExporter().export(data, path, kw)
            messagebox.showinfo(
                "Saved",
                f"✅ {len(data)} records saved to:\n{path}"
            )
            try:
                os.startfile(path)       # Windows
            except:
                try:
                    os.system(f'open "{path}"')  # macOS
                except:
                    pass
        except Exception as e:
            messagebox.showerror("Export Error", str(e))

    # ──────────────────────────────────────
    #  Clear
    # ──────────────────────────────────────
    def _clear(self):
        if not messagebox.askyesno("Clear", "Clear all results?"):
            return
        self.results = []
        self.tree.delete(*self.tree.get_children())
        self._log_clear()
        self.prog_bar.set(0)
        self.prog_lbl.configure(text="0 / 0")
        self.phase_lbl.configure(text="Phase: –")
        for w in (self.s_total, self.s_phone, self.s_email, self.s_web):
            w.configure(text="0")
        self.btn_export.configure(state="disabled")
        self._set_status("Ready", ACCENT2)

    # ──────────────────────────────────────
    #  Table helpers
    # ──────────────────────────────────────
    def _sort(self, col):
        rev = self._sort_reverse.get(col, False)
        items = [(self.tree.set(k, col), k)
                 for k in self.tree.get_children("")]
        try:
            items.sort(key=lambda x: float(x[0]) if x[0] else 0,
                       reverse=rev)
        except:
            items.sort(key=lambda x: x[0].lower(), reverse=rev)
        for i, (_, k) in enumerate(items):
            self.tree.move(k, "", i)
        self._sort_reverse[col] = not rev

    def _copy_row(self):
        sel = self.tree.selection()
        if sel:
            vals = self.tree.item(sel[0])["values"]
            self.clipboard_clear()
            self.clipboard_append("\t".join(str(v) for v in vals))

    def _del_row(self):
        for iid in list(self.tree.selection()):
            idx = int(iid) - 1
            if 0 <= idx < len(self.results):
                self.results.pop(idx)
            self.tree.delete(iid)


# ═══════════════════════════════════════════════════════
if __name__ == "__main__":
    if not SELENIUM_OK:
        print("pip install selenium webdriver-manager")
    if not OPENPYXL_OK:
        print("pip install openpyxl")
    App().mainloop()